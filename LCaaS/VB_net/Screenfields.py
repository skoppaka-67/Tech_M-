# -*- coding: utf-8 -*-
"""
Created on Tue Feb 23 16:08:44 2021

@author: naveen
"""
from os import walk
from os.path import join, isfile
import json
import re

from pymongo import MongoClient
client = MongoClient('localhost', 27017)
db = client["vb"]
col = db["screenfields"]

import openpyxl

    
with open("config.json") as f: #importing json data
    data = json.load(f)

path = data["path"]
check1 = data["check1"]
check2 = data["check2"]
check3 = data["check3"]
check4 = data["check4"]
check5 = data["check5"]
check6 = data["check6"]
check7 = data["check7"]
check8 = data["check8"]
panel = data["panel"]
extention1 = data["extention1"]
extention2 = data["extention2"]
endfunc1 = data["endfunc1"]
endfunc2 = data["endfunc2"]
pageload = data["pageload"]
modaldict = data["modaldict"]
endwithfunc = data["endwithfunc"]
screenfields = data["screenfields"]

def withlines(Lines,ID):#function to return the lines between with in with lines
    startline = 0
    endline = 0
    templines = []
    for line in Lines:
        print(Lines.index(line))
        if "With "+ID in line:
            startline = Lines.index(line)
        if startline != 0 and endwithfunc in line:          
            endline = Lines.index(line,startline)
        if startline != 0 and endline != 0:
            templines = Lines[startline+1:endline]
            templines = [line2.strip() for line2 in templines]
            startline = 0
            endline = 0  
            break                  
    return templines

def funclines(funcname,Lines):#function to return the lines of the function
    startline = 0
    endline = 0
    templines =[]
    for line in Lines:
        if funcname in line:
            startline = Lines.index(line)
        if (line.lower().__contains__(endfunc1) or line.lower().__contains__(endfunc2)) and startline!=0:
            endline = Lines.index(line)
        if startline!=0 and endline!=0:
            templines = Lines[startline+1:endline]
            
    return templines
  
def checkinternalfunc(Lines):#function to return lines of internal functions that are called in the page_load function
    templines = []
    Page_Load_Lines = funclines(pageload,Lines)
    for line in Page_Load_Lines:
        templist = line.split()
        for value in templist:
            if "()" in value and "." not in value:
                #print(value)
                templines.append(funclines(value[0:-2],Lines))             
    return templines

def getextensionfiles(path,extention):#function to return the file paths in the given directory with given extention
    filelist = []
    for r, d, f in walk(path):
        for file in f:
            if file.endswith(extention):
                filepath =join(r, file)
                filelist.append(filepath)
                
    return filelist

#def findscreenfield(SF,line):
#    if SF+"=" in line:
#        if line.split(SF+"=")[1].split(">")[0][0] == '"' and line.split(SF+"=")[1].split(">")[0][-1] == '"':
#            return line.split(SF+"=")[1].split(">")[0].replace('"', '')
#        elif line.split(SF+"=")[1].split()[0][0] == '"' and line.split(SF+"=")[1].split()[0][-1] == '"' :   
#            return line.split(SF+"=")[1].split()[0].replace('"', '')  
#        else:
#            return ""
#    else:
#        return ""
    
#def findSFinlinkfile(SF,line,line2):
#    if line.split(check3)[1].split(" ")[0].replace('"', '')+"."+ SF +" = " in line2:
#        return line2.split(line.split(check3)[1].split(" ")[0].replace('"', '')+"."+ SF +" = ")[1] 
#    else:
#        return ""
    
#def findSFinwithline(SF,line2):
#    if "."+SF in line2:
#        return line2.split("."+SF+" = ")[1] .split()[0].replace('"', '') 
#    else:
#        return "" 
    
def findscreenfield(SF,line):# function to get property in the line it can be string or variable
    print('MF',SF)
    if line.split(SF+"=")[1][0]=='"':
        return re.findall('"([^"]*)"', line.split(SF+"=")[1])[0].strip()
    elif line.split(SF+"=")[1][0]=="'":
        return re.findall("'([^']*)'", line.split(SF+"=")[1])[0].strip()    
    elif (line.split(SF+"=")[1][0]!='"' or line.split(SF+"=")[1][0]!="'") and line.split(SF+"=")[1].split()[0][-1]!='>':
        return line.split(SF+"=")[1].split()[0]
    elif (line.split(SF+"=")[1][0]!='"' or line.split(SF+"=")[1][0]!="'") and line.split(SF+"=")[1].split()[0][-1]=='>':
        return line.split(SF+"=")[1].split('>')[0] 
    else:
        return ""
    
def findSFinlinkfile(SF,ID,line2):# function to get property in the pageload and internal function lines
    print('LF',SF)
    return line2.split(ID+"."+ SF +" = ")[1].strip() 


def findSFinwithline(SF,line2):# function to get property in the with lines
    print('WF',SF)
    return line2.split("."+SF+" = ")[1] .split()[0].replace('"', '').strip()  


def checkline(check2,check5,line,Lines):#a recursive function to get full line if it not present in single line
    output = ""
    if (check2 in line) or (check5 in line):
        output = output +" "+ line
    else:
        output = output +" "+ line
        line = Lines[Lines.index(line)+1]
        output = output + checkline(check2,check5,line,Lines)
    return output

#alllist = getextensionfiles(path,extention1)
def getreport(path,extention1):
    alldict = []
    for filepath in getextensionfiles(path,extention1):# iterate through every file
        print('----------------filename :',filepath.split("\\")[-1],'----------------------')
        openFile = open(filepath,encoding="utf8")#reads file
        SectionName = []
        SectionID = []
        Lines = openFile.readlines()
        for line in Lines:#iterate through lines of file
            if (check1 in line) and ((check2 not in line) and (check5 not in line)):#a recursive function to get full line if it not present in single line
                print('======================Line is alongating=====================')
                line = checkline(check2,check5,line,Lines) 
                print(line)
            if check1 in line and (check3 or check8 in line):
                tempdict = modaldict.copy()
                tempdict["filename"] = filepath.split("\\")[-1]#filename
                tempdict["Application"] = filepath.split("\\")[-2].upper()#foldername
                #tempdict["ScreenField"] = line.split(check3)[1].split()[0].replace('"','')# ID or Screenfield
                if check3 in line:
                    tempdict["ScreenField"] = findscreenfield("ID",line)
                if check8 in line:
                    tempdict["ScreenField"] = findscreenfield("id",line)    
                print('ScreenField:', tempdict["ScreenField"])
                tempdict["Type"] = line.split(check1)[1].split(" ")[0]#tagname or type
                
                wb_obj = openpyxl.load_workbook("controls_list.xlsx")
                sheet_obj = wb_obj.active
                m_row = sheet_obj.max_row  
                for i in range(2, m_row + 1):
                    cell_obj = sheet_obj.cell(row = i, column = 2)
                    if tempdict["Type"].lower()== cell_obj.value.lower():
                        cell_obj2 = sheet_obj.cell(row = i, column = 3)
                        tempdict["Attributes"] = cell_obj2.value                  
                    
                for key , value in screenfields.items():#getting all the screenfields present in the asp line
                    if value+"=" in line:
                        tempdict[key] = findscreenfield(value,line)                     
                    
                if tempdict["Type"].lower().__contains__(panel):#these 3 conditions to get sectionname and sectionid fields
                    SectionName.append(tempdict["Type"])
                    SectionID.append(tempdict["ScreenField"])
                if len(SectionName)!=0:
                    tempdict["SectionName"] = SectionName[-1]
                if len(SectionID)!=0:
                    tempdict["SectionID"] = SectionID[-1]
                    
                if tempdict["Type"] == "RequiredFieldValidator":
                    if "ErrorMessage"+"=" in line:
                        tempdict["ErrorMessage"] = findscreenfield("ErrorMessage",line)
                    if "ControlToValidate"+"=" in line:
                        tempdict["ControlToValidate"] = findscreenfield("ControlToValidate",line)
                if tempdict["Type"] == "CompareValidator":
                    if "ErrorMessage"+"=" in line:
                        tempdict["ErrorMessage"] = findscreenfield("ErrorMessage",line) 
                    if "ControlToValidate"+"=" in line:
                        tempdict["ControlToValidate"] = findscreenfield("ControlToValidate",line)
                    if "ControlToCompare"+"=" in line:
                        tempdict["ControlToCompare"] = findscreenfield("ControlToCompare",line)
                if tempdict["Type"] == "RangeValidator":
                    if "ErrorMessage"+"=" in line:
                        tempdict["ErrorMessage"] = findscreenfield("ErrorMessage",line) 
                    if "ControlToValidate"+"=" in line:
                        tempdict["ControlToValidate"] = findscreenfield("ControlToValidate",line)
                    if "MaximumValue"+"=" in line:
                        tempdict["MaximumValue"] = findscreenfield("MaximumValue",line)
                    if "MinimumValue"+"=" in line:
                        tempdict["MinimumValue"] = findscreenfield("MinimumValue",line)
                if tempdict["Type"] == "RegularExpressionValidator":
                    if "ErrorMessage"+"=" in line:
                        tempdict["ErrorMessage"] = findscreenfield("ErrorMessage",line)
                    if "ControlToValidate"+"=" in line:
                        tempdict["ControlToValidate"] = findscreenfield("ControlToValidate",line)
                    if "ValidationExpression"+"=" in line:
                        tempdict["ValidationExpression"] = findscreenfield("ValidationExpression",line)
                if tempdict["Type"] == "TextBox":
                    if "TextMode"+"=" in line:
                        tempdict["TextMode"] = findscreenfield("TextMode",line)  
                
                #get the vb files which are related to  this particular aspx file                   
                filepaths = getextensionfiles(path,filepath.split("\\")[-1]+extention2)+getextensionfiles(path,filepath.split("\\")[-1].split(".")[0]+extention2)
                for linkfile in filepaths:                    
                    if isfile(linkfile):
                        openFile2 = open(linkfile,encoding="utf8")
                        Lines2 = openFile2.readlines()
                        startline = 0
                        endline = 0
                        templines = 0
                        for line2 in funclines(pageload,Lines2)+checkinternalfunc(Lines2):                           
                            if "With "+tempdict["ScreenField"] in line2:#these below 3 if conditions to get lines of with format
                                startline = Lines2.index(line2)
                            if startline!=0 and endwithfunc in line2:                               
                                endline = Lines2.index(line2)
                            if startline!=0 and endline!=0:                               
                                templines = Lines2[startline:endline]
                            
                            for key , value in screenfields.items():#getting all the screenfields present in the pageload function and internalfunctions called
                                if tempdict["ScreenField"]+"."+ value +" = " in line2: 
                                    tempdict[key] = findSFinlinkfile(value,tempdict["ScreenField"],line2)                                
                            
                                
                            if tempdict["Type"] == "TextBox":
                                if tempdict["ScreenField"]+"."+ "TextMode" +" = " in line2:
                                    tempdict["TextMode"] = findSFinlinkfile("TextMode",tempdict["ScreenField"],line2)
                            if tempdict["Type"] == "DropDownList": 
                                if tempdict["ScreenField"]+"."+ "DataSource" +" = " in line2:
                                    tempdict["DataSource"] = findSFinlinkfile("DataSource",tempdict["ScreenField"],line2) 
                                if tempdict["ScreenField"]+"."+ "DataTextField" +" = " in line2:
                                    tempdict["DataTextField"] = findSFinlinkfile("DataTextField",tempdict["ScreenField"],line2)
                                if tempdict["ScreenField"]+"."+ "DataValueField" +" = " in line2:
                                    tempdict["DataValueField"] = findSFinlinkfile("DataValueField",tempdict["ScreenField"],line2)
                               
                        #print(templines)
                        if templines!=0:  
                            templine = " ".join(templines)
                            #for line2 in templines:
                            for key , value in screenfields.items():#getting all the screenfields present in the with format lines
                                if "."+ value in templine:
                                    tempdict[key] = findSFinwithline(value,templine)        
                                    

                            if tempdict["Type"] == "TextBox":
                                if "."+ "TextMode" in templine:
                                    tempdict["TextMode"] = findSFinwithline("TextMode",templine)                                
                            if tempdict["Type"] == "DropDownList":
                                if "."+ "DataSource" in templine:
                                    tempdict["DataSource"] = findSFinwithline("DataSource",templine)  
                                if "."+ "DataTextField" in templine:
                                    tempdict["DataTextField"] = findSFinwithline("DataTextField",templine)
                                if "."+ "DataValueField" in templine:
                                    tempdict["DataValueField"] = findSFinwithline("DataValueField",templine)
    
                alldict.append(tempdict.copy())
            elif check4 in line and line.lower().__contains__(panel):#delete the sectionname and sectionid after the block of code completed
                SectionName.pop()
                SectionID.pop()

    return alldict                        

def insertfunc():
    if col.count() != 0 :
        col.drop()
        col.insert_many(getreport(path,extention1))
        col.insert_one({"type" : "metadata",
                    "headers" : ["filename","Application", "SectionID", "SectionName", "ScreenField", 
                                "Type", "TextMode", "Enable", "Tooltip", "Visible","DataTextField", 
                                "DataValueField" ]})
    else:
        col.insert_many(getreport(path,extention1))
        col.insert({"type" : "metadata",
                    "headers" : ["filename","Application", "SectionID", "SectionName", "ScreenField", 
                                "Type", "TextMode", "Enable", "Tooltip", "Visible","DataTextField", 
                                "DataValueField" ]})       
json.dump(getreport(path,extention1), open('output.json', 'w'), indent=4) 
insertfunc()


    