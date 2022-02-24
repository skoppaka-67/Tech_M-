from openpyxl import Workbook, load_workbook
from pymongo import MongoClient
import timeit
import logging
import os


my_path = os.path.abspath(os.path.dirname(__file__))

#Logger settings
logging.basicConfig(filename='app.log', filemode='a', format='[%(levelname)s]%(asctime)s:  %(message)s',level=logging.DEBUG)
logger = logging.getLogger(__name__)
########################### EXCEL CONFIGURATION #################################
# Specify the folder where the workbook is present
#file_location = "E:\\Work\\Work\\Automation\\Projects\\Mainframe Static Code Analyser\\Naga Inputs\\Sample_Code\\"
file_location = "D:\\CRST_FULL\\"
# Specify the file name
file_name = "CRST_appname_cobol.xlsx"
# Specify the sheet name that contains the data
sheet_name = "Sheet1"
#path = os.path.join(file_location)
########################### MONGODB CONFIGURATION #################################
hostname = 'localhost'
port = 27017
database_name = 'CRST_FULLV1'

# Excel Initialization
wb = load_workbook(file_location + file_name, data_only=True)
sh = wb[sheet_name]

# MongoDB Initialization
client = MongoClient(hostname, port)
db = client[database_name]

# def updateMasterInventoryAPPNAME():
#     time_update_minv_start=timeit.default_timer()
#     # Updating master inventory
#     try:
#         for ite in range(2, sh.max_row):
#             component_name = sh.cell(row=ite, column=1).value
#             component_type = sh.cell(row=ite, column=2).value
#             application_name = sh.cell(row=ite, column=3).value
#             # if db.master_inventory_report.update_one({"component_name": component_name},
#             #                                          {"$set": {"application": application_name}}).acknowledged:
#             #     sh.cell(row=ite, column=4).value = 'UPDATED'
#             #     print('MINV =====> Updated', sh.cell(row=ite, column=1).value + '\'s application as',
#             #           sh.cell(row=ite, column=3).value)
#             # else:
#             #     sh.cell(row=ite, column=4).value = 'NOT UPDATED'
#             row = db.master_inventory_report.update_one({"component_name": component_name},{"$set": {"application": application_name}})
#             print(row)
#         time_update_minv_stop=timeit.default_timer()
#         time_taken = time_update_minv_stop - time_update_minv_start
#         print('Time taken :',time_taken)
#     except TypeError as e:
#         print(e)

def updateCrossReferenceAPPNAME_CALLING():
    # Updating calling_app_name

    try:
     for ite in range(2, sh.max_row):
        print(sh.cell(row=ite, column=1).value, sh.cell(row=ite, column=3).value)
        component_name = sh.cell(row=ite, column=1).value
        component_type = sh.cell(row=ite, column=2).value
        application = sh.cell(row=ite, column=3).value
        if component_type == "COBOL":
            component_name = component_name.split(".")[0]
        elif component_type == "COPYBOOK":
            component_name = component_name
        elif component_type == "APS":
            component_name = component_name.split(".")[0]
        elif component_type == "BMS":
            component_name = component_name.split(".")[0]
        elif component_type == "JCL":
            component_name = component_name.split(".")[0]
        elif component_type == "PROC":
            component_name = component_name.split(".")[0]
        elif component_type == "SYSIN":
            component_name = component_name.split(".")[0]

        if db.cross_reference_report.update_many({"$and":[{"component_name": component_name},{"component_type": component_type}]}, {
            "$set": {"calling_app_name": application.rstrip()}}).acknowledged:
            #sh.cell(row=ite, column=5).value = 'UPDATED'
            print('CALLING_APP =====>Updated', sh.cell(row=ite, column=1).value + '\'s application as',
                  sh.cell(row=ite, column=3).value)
        else:
            sh.cell(row=ite, column=5).value = 'NOT UPDATED'
    except TypeError as e:
        print(e)
def updateCrossReferenceAPPNAME_CALLED():
    # Updating calling_app_name
    try:
        for ite in range(2, sh.max_row):
            print(sh.cell(row=ite, column=1).value, sh.cell(row=ite, column=3).value)
            component_name = sh.cell(row=ite, column=1).value
            component_type = sh.cell(row=ite, column=2).value
            application = sh.cell(row=ite, column=3).value
            if component_type == "COBOL":
                component_name = component_name.split(".")[0]
            elif component_type == "COPYBOOK":
                component_name = component_name
            elif component_type == "APS":
                component_name = component_name.split(".")[0]
            elif component_type == "BMS":
                component_name = component_name.split(".")[0]
            elif component_type == "JCL":
                component_name = component_name.split(".")[0]
            elif component_type == "PROC":
                component_name = component_name.split(".")[0]
            elif component_type == "SYSIN":
                component_name = component_name.split(".")[0]

            if db.cross_reference_report.update_many({"$and":[{"called_name": component_name},{"called_type": component_type}]}, {
                "$set": {"called_app_name": application.rstrip()}}).acknowledged:
                #sh.cell(row=ite, column=6).value = 'UPDATED'
                print('CALLED_APP =====> Updated', sh.cell(row=ite, column=1).value + '\'s application as',
                      sh.cell(row=ite, column=3).value)
            else:
                sh.cell(row=ite, column=5).value = 'NOT UPDATED'
    except TypeError as e:
        print(e)


def updateMasterInventoryAPPNAME():
    time_update_minv_start=timeit.default_timer()
    # Updating master inventory
    try:
        for ite in range(2, sh.max_row):
            component_name = sh.cell(row=ite, column=1).value
            component_type = sh.cell(row=ite, column=2).value
            application_name = sh.cell(row=ite, column=3).value
            # if db.master_inventory_report.update_one({"component_name": component_name},
            #                                          {"$set": {"application": application_name}}).acknowledged:
            #     sh.cell(row=ite, column=4).value = 'UPDATED'
            #     print('MINV =====> Updated', sh.cell(row=ite, column=1).value + '\'s application as',
            #           sh.cell(row=ite, column=3).value)
            # else:
            #     sh.cell(row=ite, column=4).value = 'NOT UPDATED'
            if component_type == "COBOL":
                component_name = component_name
            elif component_type == "COPYBOOK":
                component_name = component_name + ".cpy"
            elif component_type == "APS":
                component_name = component_name
            elif component_type == "BMS":
                component_name = component_name
            elif component_type == "JCL":
                component_name = component_name
            elif component_type == "PROC":
                component_name = component_name
            elif component_type == "SYSIN":
                component_name = component_name
            row = db.master_inventory_report.update_one({"$and":[{"component_name": component_name},{"component_type": component_type}]},{"$set": {"application": application_name.rstrip()}})
            print(row)
        time_update_minv_stop=timeit.default_timer()
        time_taken = time_update_minv_stop - time_update_minv_start
        print('Time taken :',time_taken)
    except TypeError as e:
        print(e)


# def updateCrossReferenceAPPNAME_CALLED():
#     # Updating calling_app_name
#     try:
#         for ite in range(2, sh.max_row):
#             print(sh.cell(row=ite, column=1).value, sh.cell(row=ite, column=3).value)
#             if db.cross_reference_report.update_many({"called_name": sh.cell(row=ite, column=1).value}, {
#                 "$set": {"called_app_name": sh.cell(row=ite, column=3).value}}).acknowledged:
#                 #sh.cell(row=ite, column=6).value = 'UPDATED'
#                 print('CALLED_APP =====> Updated', sh.cell(row=ite, column=1).value + '\'s application as',
#                       sh.cell(row=ite, column=3).value)
#             else:
#                 sh.cell(row=ite, column=5).value = 'NOT UPDATED'
#     except TypeError as e:
#         print(e)


def updateCICSAPPNAME():
    time_update_minv_start=timeit.default_timer()
    # Updating master inventory
    try:
        for ite in range(2, sh.max_row):
            component_name = sh.cell(row=ite, column=1).value
            component_type = sh.cell(row=ite, column=2).value
            application_name = sh.cell(row=ite, column=3).value
            # if db.master_inventory_report.update_one({"component_name": component_name},
            #                                          {"$set": {"application": application_name}}).acknowledged:
            #     sh.cell(row=ite, column=4).value = 'UPDATED'
            #     print('MINV =====> Updated', sh.cell(row=ite, column=1).value + '\'s application as',
            #           sh.cell(row=ite, column=3).value)
            # else:
            #     sh.cell(row=ite, column=4).value = 'NOT UPDATED'

            row = db.cics_field.update_many({"bms_name": component_name.split(".")[0]},{"$set": {"application": application_name.rstrip()}})
            print(component_name.strip("."))
        time_update_minv_stop=timeit.default_timer()
        time_taken = time_update_minv_stop - time_update_minv_start
        print('Time taken :',time_taken)
    except TypeError as e:
        print(e)

#
# def updateCrossReferenceAPPNAME_CALLED():
#     # Updating calling_app_name
#     try:
#         for ite in range(2, sh.max_row):
#             print(sh.cell(row=ite, column=1).value, sh.cell(row=ite, column=3).value)
#             if db.cross_reference_report.update_many({"called_name": sh.cell(row=ite, column=1).value}, {
#                 "$set": {"called_app_name": sh.cell(row=ite, column=3).value}}).acknowledged:
#                 #sh.cell(row=ite, column=6).value = 'UPDATED'
#                 print('CALLED_APP =====> Updated', sh.cell(row=ite, column=1).value + '\'s application as',
#                       sh.cell(row=ite, column=3).value)
#             else:
#                 sh.cell(row=ite, column=5).value = 'NOT UPDATED'
#     except TypeError as e:
#         print(e)

def updateCrudAPPNAME_CALLING():
    # Updating calling_app_name\
    try:
        for ite in range(2, sh.max_row):
            print(sh.cell(row=ite, column=1).value, sh.cell(row=ite, column=3).value)
            component_name = sh.cell(row=ite, column=1).value
            component_type = sh.cell(row=ite, column=2).value
            application = sh.cell(row=ite, column=3).value
            if component_type == "COBOL":
                component_name = component_name
            elif component_type == "COPYBOOK":
                component_name = component_name
            elif component_type == "APS":
                component_name = component_name
            elif component_type == "BMS":
                component_name = component_name
            elif component_type == "JCL":
                component_name = component_name
            elif component_type == "PROC":
                component_name = component_name
            elif component_type == "SYSIN":
                component_name = component_name
            if db.crud_report.update_many({"$and":[{"component_name": component_name},{"component_type": component_type}]}, {
                "$set": {"calling_app_name": application.rstrip()}}).acknowledged:
                #sh.cell(row=ite, column=5).value = 'UPDATED'
                print('CALLING_APP =====>Updated', sh.cell(row=ite, column=1).value + '\'s application as',
                      sh.cell(row=ite, column=3).value)
            else:
                sh.cell(row=ite, column=5).value = 'NOT UPDATED'

    except TypeError as e :
        print( e, sh.cell(row=ite, column=3).value)
def save():
    wb.save(file_location+file_name)



def updateBRE():
    time_update_minv_start=timeit.default_timer()
    # Updating master inventory
    try:
        for ite in range(2, sh.max_row):
            component_name = sh.cell(row=ite, column=1).value
            component_type = sh.cell(row=ite, column=2).value
            application_name = sh.cell(row=ite, column=3).value
            if component_type != "COBOL":
                continue
            else:
                component_name = component_name[:-3]
            # component_name = component_name

            row = db.bre_rules_report.update_many({"pgm_name": component_name},{"$set": {"application": application_name.rstrip()}})
            print(component_name,application_name)
        time_update_minv_stop=timeit.default_timer()
        time_taken = time_update_minv_stop - time_update_minv_start
        print('Time taken :',time_taken)
    except TypeError as e:
        print(e)



def update_orphan():
    time_update_minv_start=timeit.default_timer()
    # Updating master inventory
    try:
        for ite in range(2, sh.max_row):
            component_name = sh.cell(row=ite, column=1).value
            component_type = sh.cell(row=ite, column=2).value
            application_name = sh.cell(row=ite, column=3).value
            if component_type == "COBOL":
                component_name = component_name.split(".")[0]
            elif component_type == "COPYBOOK":
                component_name = component_name
            elif component_type == "APS":
                component_name = component_name.split(".")[0]
            elif component_type == "BMS":
                component_name = component_name.split(".")[0]
            elif component_type == "JCL":
                component_name = component_name.split(".")[0]
            elif component_type == "PROC":
                component_name = component_name.split(".")[0]
            elif component_type == "SYSIN":
                component_name = component_name.split(".")[0]

            row = db.orphan_report.update_many({"$and":[{"component_name": component_name},{"component_type": component_type}]},{"$set": {"application": application_name.rstrip()}})
            print(component_name,application_name)
        time_update_minv_stop=timeit.default_timer()
        time_taken = time_update_minv_stop - time_update_minv_start
        print('Time taken :',time_taken)
    except TypeError as e:
        print(e)


def update_dropImpact():
    time_update_minv_start=timeit.default_timer()
    # Updating master inventory
    try:
        for ite in range(2, sh.max_row):
            component_name = sh.cell(row=ite, column=1).value
            component_type = sh.cell(row=ite, column=2).value
            application_name = sh.cell(row=ite, column=3).value
            if component_type == "COBOL":
                component_name = component_name.split(".")[0]
            elif component_type == "COPYBOOK":
                component_name = component_name
            elif component_type == "APS":
                component_name = component_name
            elif component_type == "BMS":
                component_name = component_name
            elif component_type == "JCL":
                component_name = component_name
            elif component_type == "PROC":
                component_name = component_name
            elif component_type == "SYSIN":
                component_name = component_name

            row = db.drop_impact.update_many({"$and":[{"drop_impact_name": component_name},{"drop_impact_type": component_type}]},{"$set": {"application": application_name.rstrip()}})
            print(component_name,application_name)
        time_update_minv_stop=timeit.default_timer()
        time_taken = time_update_minv_stop - time_update_minv_start
        print('Time taken :',time_taken)
    except TypeError as e:
        print(e)

def update_commented_lines():


    time_update_minv_start=timeit.default_timer()
    # Updating master inventory
    try:
        for ite in range(2, sh.max_row):
            component_name = sh.cell(row=ite, column=1).value
            component_type = sh.cell(row=ite, column=2).value
            application_name = sh.cell(row=ite, column=3).value
            if component_type == "COBOL":
                component_name = component_name
            elif component_type == "COPYBOOK":
                component_name = component_name
            elif component_type == "APS":
                component_name = component_name
            elif component_type == "BMS":
                component_name = component_name
            elif component_type == "JCL":
                component_name = component_name
            elif component_type == "PROC":
                component_name = component_name
            elif component_type == "SYSIN":
                component_name = component_name

            row = db.cobol_output.update_many({"$and":[{"component_name": component_name},{"component_type": component_type}]},{"$set": {"application": application_name.rstrip()}})
            print(component_name,application_name)
        time_update_minv_stop=timeit.default_timer()
        time_taken = time_update_minv_stop - time_update_minv_start
        print('Time taken :',time_taken)
    except TypeError as e:
        print(e)




#
# updateMasterInventoryAPPNAME()

# updateCrossReferenceAPPNAME_CALLING()
# updateCrossReferenceAPPNAME_CALLED()
# updateCrudAPPNAME_CALLING()
update_commented_lines()
# updateBRE()
# updateCICSAPPNAME()
# update_dropImpact()
# update_orphan()


save()
