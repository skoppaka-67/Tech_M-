from openpyxl import Workbook, load_workbook
from pymongo import MongoClient
import timeit
import logging
import os


my_path = os.path.abspath(os.path.dirname(__file__))

#Logger settings
logging.basicConfig(filename='app.log', filemode='a', format='[%(levelname)s]%(asctime)s:  %(message)s',level=logging.DEBUG)
logger = logging.getLogger(__name__)
########################### EXCEL CONFIGURATION #################################
# Specify the folder where the workbook is present
#file_location = "E:\\Work\\Work\\Automation\\Projects\\Mainframe Static Code Analyser\\Naga Inputs\\Sample_Code\\"
file_location = "D:\\POC\\"
# Specify the file name
file_name = "appname_update.xlsx"
# Specify the sheet name that contains the data
sheet_name = "Sheet1"
#path = os.path.join(file_location)
########################### MONGODB CONFIGURATION #################################
hostname = 'localhost'
port = 27017
database_name = 'TW'

# Excel Initialization
wb = load_workbook(file_location + file_name, data_only=True)
sh = wb[sheet_name]

# MongoDB Initialization
client = MongoClient(hostname, port)
db = client[database_name]

def updateMasterInventoryAPPNAME():
    time_update_minv_start=timeit.default_timer()
    # Updating master inventory
    for ite in range(2, sh.max_row):
        component_name = sh.cell(row=ite, column=1).value
        component_type = sh.cell(row=ite, column=2).value
        application_name = sh.cell(row=ite, column=3).value
        # if db.master_inventory_report.update_one({"component_name": component_name},
        #                                          {"$set": {"application": application_name}}).acknowledged:
        #     sh.cell(row=ite, column=4).value = 'UPDATED'
        #     print('MINV =====> Updated', sh.cell(row=ite, column=1).value + '\'s application as',
        #           sh.cell(row=ite, column=3).value)
        # else:
        #     sh.cell(row=ite, column=4).value = 'NOT UPDATED'
        row = db.master_inventory_report.update_one({"component_name": component_name},{"$set": {"application": application_name}})
        # print(row)
    time_update_minv_stop=timeit.default_timer()
    time_taken = time_update_minv_stop - time_update_minv_start
    print('Time taken :',time_taken)


def updateCrossReferenceAPPNAME_CALLING():
    # Updating calling_app_name
    for ite in range(2, sh.max_row):
        print(sh.cell(row=ite, column=1).value, sh.cell(row=ite, column=3).value)
        if db.cross_reference_report.update_many({"component_name": sh.cell(row=ite, column=1).value}, {
            "$set": {"calling_app_name": sh.cell(row=ite, column=3).value}}).acknowledged:
            #sh.cell(row=ite, column=5).value = 'UPDATED'
            print('CALLING_APP =====>Updated', sh.cell(row=ite, column=1).value + '\'s application as',
                  sh.cell(row=ite, column=3).value)
        else:
            sh.cell(row=ite, column=5).value = 'NOT UPDATED'

def updateCrossReferenceAPPNAME_CALLED():
    # Updating calling_app_name
    for ite in range(2, sh.max_row):
        print(sh.cell(row=ite, column=1).value, sh.cell(row=ite, column=3).value)
        if db.cross_reference_report.update_many({"called_name": sh.cell(row=ite, column=1).value}, {
            "$set": {"called_app_name": sh.cell(row=ite, column=3).value}}).acknowledged:
            #sh.cell(row=ite, column=6).value = 'UPDATED'
            print('CALLED_APP =====> Updated', sh.cell(row=ite, column=1).value + '\'s application as',
                  sh.cell(row=ite, column=3).value)
        else:
            sh.cell(row=ite, column=5).value = 'NOT UPDATED'


def save():
    wb.save(file_location+file_name)

updateMasterInventoryAPPNAME()
updateCrossReferenceAPPNAME_CALLING()
updateCrossReferenceAPPNAME_CALLED()
save()
