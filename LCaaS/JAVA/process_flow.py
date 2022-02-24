# -*- coding: utf-8 -*-
"""
@author: KN00636678

JAVA Process flow report

Dependencies: 
config file with filepath, collectionname and dbname
componenttype collection name
Run the componenttype.py file before using using report

"""
import os
import config
import copy
import re
from bs4 import BeautifulSoup
import openpyxl
import json
import pandas as pd

from pymongo import MongoClient
client = MongoClient('localhost', 27017)

filespath = config.filespath
dbname = config.dbname
collectionname = config.processflowcn
componentcollection = config.componenttypecn
col2 = client[dbname][componentcollection]

remainingfilepath = filespath + "\\Property_codedump\\src\\"

modaljson = {"component_name": "",
             "event_name": "",
             "application":"",
             "nodes": [],
             "links": []}

extensions = ['.java', '.jsp', '.js', '.css']

tags = ['form', 'a']

csyntax = {".jsp":('<!--','-->','<%--','--%>'),#different comment syntax for different files
           ".css":('//','','/*','*/'),
           ".java":('//','','/*','*/'),
           ".js"  :('//','','/*','*/')}

def getallfiles(filespath,extensions):#function to return the file paths in the given directory with given extention
    filelist = []
    for root, dirs, files in os.walk(filespath):
        for file in files:
            if file.lower().endswith(tuple([item.lower() for item in extensions])):
                filelist.append(os.path.join(root, file))
    return filelist

def removecommentlines(Lines,extension):
    start1,end1,start2,end2 = csyntax["."+extension.lower()]
    Lines = [line for line in Lines if not line.strip().startswith(start1)]
    commentflag = False
    for index, line in enumerate(Lines):#iterate through lines of file                      
        if commentflag:#between lines of multicomment line
            Lines[index] = ''
        if start2 in line:  
            Lines[index] = line.split(start2)[0].strip()                
            commentflag = True
        if end2 in line:#end of multicomment line
            Lines[index] = line.split(end2)[1].strip()
            commentflag = False                         
      
    Lines = [line for line in Lines if line.strip()!=""]    
    return Lines

def getdomethod(eventtag):
    if "method" in eventtag.keys():
        if eventtag["method"].lower() == "get":
            return "doGet"
        elif eventtag["method"].lower() == "post":
            return "doPost"                                    
    else:
        return "doGet"    

def geteventtriggers(filespath):
    output = []
    output2 = []
    cursor = list(col2.find({'type': {"$ne": "metadata"}},{"component_name":1,"application":1,"WebServlet":1,"_id":0}))
    cursor = [item for item in cursor if item["WebServlet"]!=None]
    for filepath in getallfiles(filespath,[".jsp"]):
        html  = open(filepath).read()#reads file
        if html.strip()!="":
            soup = BeautifulSoup(html,features="lxml")
            for tag in tags:
                listofsingletag = soup.find_all(tag) 
                # print(listofsingletag)
                for singletag in listofsingletag: 
                    eventtag = singletag.attrs
                    if tag == "form" and "action" in eventtag.keys():
                        filename = eventtag["action"]#+".java"
                        filename = filename.split("/")[-1]
                        # print(filename)
                        if filename in [item["WebServlet"].split("/")[-1] for item in cursor]:                            
                            tempdict = {}
                            tempdict["component_name"] = filepath.split("\\")[-1]
                            tempdict["called_name"] = [item["component_name"] for item in cursor if item["WebServlet"].split("/")[-1]==filename][0]
                            tempdict["application"] = [item["application"] for item in cursor if item["component_name"]==tempdict["called_name"]][0]
                            tempdict["eventname"] = getdomethod(eventtag)    
                            output.append(tempdict.copy())  
                        elif filename.count(".")==0 and filename!="" and not filename.startswith("#"):
                            # print(filename)
                            eventname = getdomethod(eventtag)                              
                            output2.append({"called_name":filename+".java","eventname":eventname})
                    elif tag == "a" and "href" in eventtag.keys():
                        filename = eventtag["href"]#+".java" 
                        filename = filename.split("/")[-1]
                        if filename in [item["WebServlet"].split("/")[-1] for item in cursor]:
                            tempdict = {}
                            tempdict["component_name"] = filepath.split("\\")[-1]
                            tempdict["called_name"] = [item["component_name"] for item in cursor if item["WebServlet"].split("/")[-1]==filename][0]
                            tempdict["application"] = [item["application"] for item in cursor if item["component_name"]==tempdict["called_name"]][0]
                            tempdict["eventname"] = "doGet"    
                            output.append(tempdict.copy())
                        elif filename.count(".")==0 and filename!="" and not filename.startswith("#"):
                            # print(filename)
                            eventname = "doGet"                               
                            output2.append({"called_name":filename+".java","eventname":eventname})                          

    # print(output2)                        
    output = [i for n, i in enumerate(output) if i not in output[n + 1:]] 
    output = [i for n, i in enumerate(output) if not (i["called_name"]  in [item["called_name"] for item in output[:n]] and i["eventname"]  in [item["eventname"] for item in output[:n]])]                     
    return output,output2

def getnoa(line):
    # print(line)
    if getparenstring(line)[0].strip()=="":
        return 0
    else:
        return getparenstring(line)[0].count(",")+1
    
def getarguments(line):
    if line.split("(")[1].split(")")[0].strip()=="":
        return [],[]
    else:
        output1 = []
        output2 = []
        arglist = line.split("(")[1].split(")")[0].split(",")
        for item in arglist:
            output1.append(item.split()[0].strip())
            output2.append(item.split()[-1].strip())
        return output1,output2

def getfirstline(Lines,index):
    line = ""
    if "(" in Lines[index] and  ")" in Lines[index]:
        return Lines[index]
    if not ")" in Lines[index]:
        line = line+Lines[index]
        return line+getfirstline(Lines,index+1)
    if not "(" in Lines[index] and ")" in Lines[index]:
        return line+Lines[index]

def getfunctionreturn(line):
    if len(line.split(" ",1)[1].split("(")[0].rsplit(' ', 1))>1:
        return line.split(" ",1)[1].split("(")[0].rsplit(' ', 1)[0]
    else:
        return ""

def getvariables(line):
    output = {}
    line = line.strip()
    if line.startswith(("private","public")):
        line = line.replace("private","").replace("public","")
    line = line.split(";")[0].split("=")[0].strip()
    variables = line.split()[-1]
    for item in variables.split(","):
        output[item]=line.split()[-2]
    return output

def getparenstring(line):
    output = []
    flag = False
    counter = 0
    for index,i in enumerate(line):
        if flag and i=="(":
            counter = counter + 1
        if flag and i==")":
            if counter!=0:
                counter = counter - 1
            else:
                flag = False
                counter = 0
                output.append(line[start+1:index])
        if i=="(" and counter==0:
            flag = True
            start = index
    return output 

def getbuiltindatatypes(file):
    output = []
    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row  
    for i in range(2, m_row + 1):
        cell_obj1 = sheet_obj.cell(row = i, column = 1)
        cell_obj2 = sheet_obj.cell(row = i, column = 2)
        cell_obj3 = sheet_obj.cell(row = i, column = 3)
        output.append({"argumenttype":cell_obj1.value,
                       "methodname"  :cell_obj2.value,
                       "returntype"  :cell_obj3.value})
    return output    

def argumentcheck(objtypes,inputarguments,checkarguments):
    builtindatatypes = getbuiltindatatypes("functionreturntype.xlsx")
    fixdatatypes = {"parseInt":"int",
                    "getInt":"int",
                    "parseFloat":"float",
                    "getFloat":"float",
                    "getString":"String"}
    if checkarguments!=['']:
        for index,item in enumerate(checkarguments):
            # item=item.strip()
            if "." in item:
                performarg = item.split(".")[0]
                methodname = item.split(".",1)[1].split("(")[0]
                if methodname in fixdatatypes.keys():
                    returntype = fixdatatypes[methodname]
                else:
                    performargtype = objtypes[performarg]
                    # print(performargtype,methodname,builtindatatypes)
                    returntype = [doc for doc in builtindatatypes if performargtype==doc["argumenttype"] and methodname==doc["methodname"]][0]["returntype"]
                # if performarg in objtypes.keys():
                #     performargtype = objtypes[performarg]  
                #     print(performargtype,methodname,builtindatatypes)
                #     returntype = [doc for doc in builtindatatypes if performargtype==doc["argumenttype"] and methodname==doc["methodname"]][0]["returntype"]
                # else:
                #     # print(performarg,methodname)
                #     returntype = [doc for doc in builtindatatypes if methodname==doc["methodname"]][0]["returntype"]
            else:
                # print("data",item)
                returntype = objtypes[item]
            if returntype==inputarguments[index]:
                continue
            else:
                return False
                break
        return True
    else:
        return True

def getfilefunctionlines(Lines):
    output = [] 
    variables = {}
    flag = False
    functionflag = False
    functionflag2 = False
    for index,line in enumerate(Lines):
        # if (line.strip().startswith("public") or line.strip().startswith("private") or line.strip().startswith("protected")) and line.strip().endswith(";"):
        #     print(line)
        if flag:
            if (line.strip().startswith("public") or line.strip().startswith("private") or line.strip().startswith("protected")) and ("(" in line) and (not line.strip().endswith(";")):
                line2 = getfirstline(Lines,index)
                value1,value2 = getarguments(line2)
                
                dict1 = {"functionname"   :line2.split("(")[0].split()[-1],
                         "noa"            :getnoa(line2),
                         "functiontype"   :line2.strip().split()[0],
                         "inputformat"    :"("+line2.split("(")[1].split(")")[0].strip()+")",
                         "functionreturn" :getfunctionreturn(line2),
                         "arguments"      :value1,
                         "argumentvalues" :value2,
                         "label"          :(line2.split(")")[0]+")").strip()}
                
                functionlines = []
                functioncheck = []
                functionflag = True

            if functionflag2:
                if line.__contains__("{"):
                    functioncheck.extend([1]*line.count("{"))
                    # functioncheck.append(1)
                
                if line.__contains__("}"):
                    if len(functioncheck)!=0:
                        for i in range(line.count("}")):
                            functioncheck.pop()
                    else:
                        functionflag2 = False
                        functionflag = False
                        functionlines.append(line.split("}",1)[0])
                        dict1["functionlines"] = functionlines
                        output.append(copy.deepcopy(dict1))

            if functionflag2:
                functionlines.append(line)                        
                 
            if functionflag and not functionflag2:
                if line.__contains__("{"):
                    functionflag2 = True
                    functionlines.append(line.split("{",1)[1])
                         
        if line.strip().startswith("public interface") or line.strip().startswith("public class"):
            flag = True
        if flag and not functionflag and not line.strip().startswith("@") and all(ele not in line for ele in ["(",")","{","}"]):
            variables.update(getvariables(line))
            # print(index,line)  
    
    for item in output:
        item["variables"] = variables
    return output 

def getallfilefunctions(filespath):
    finaloutput = {}
    for filepath in getallfiles(filespath,[".java"]):# iterate through every file
        # print(filepath.split("\\")[-1])
        Lines = open(filepath).readlines()
        Lines = removecommentlines(Lines,"java")
        # if filepath.split("\\")[-1] in finaloutput.keys():
        #     print("same file repeated",filepath.split("\\")[-1])
        finaloutput[filepath.split("\\")[-1]]= getfilefunctionlines(Lines)
    return finaloutput

def getfunctioncalls(functioninfile,allobjects,importfilefunctions):
    functionlines = functioninfile["functionlines"]
    filename = functioninfile["component_name"]
    objtypes = {}
    for index,obj in enumerate(functioninfile["argumentvalues"]):
        objtypes[obj]=functioninfile["arguments"][index]
    for line in functionlines:
        if all(ele in line for ele in [" new ","=","()"]) and line.strip().endswith(";"):
            objtypes[line.strip().split("=",1)[0].split()[-1].strip()]=line.strip().split("(")[0].rsplit("new",1)[1].strip()
        elif "=" in line and line.strip().endswith(";") and len(line.strip().split("=")[0].split())>1 and all(ele not in line.strip().split("=")[0] for ele in ["+","."]):
            # print(line)
            objtypes.update(getvariables(line))
    objtypes.update(functioninfile["variables"])        
    # print(objtypes)
    output = []
    for index,line in enumerate(functionlines):
        nameslist = re.findall('"([^"]*)"', line)
        for name in nameslist:
            if name.lower().endswith(".jsp"):
                output.append({"functionname":name.split("//")[-1].split("\\")[-1].split("/")[-1],"component_name":filename})
        for function in importfilefunctions:
            if function["functionname"]+"(" in line and not "new "+function["functionname"]+"(" in line and function["noa"]==getnoa(line.split(function["functionname"],1)[1]):
                arguments =  getparenstring(line.split(function["functionname"],1)[1])[0].split(",")
                arguments = [arg.strip() for arg in arguments]
                # print(line)
                # print("argcheck",argumentcheck(objtypes,function["arguments"],arguments))
                if "."+function["functionname"]+"(" in line:
                    obj = line.split("."+function["functionname"]+"(")[0].split()[-1].split(",")[-1].split("(")[-1].split("=")[-1]
                    if obj in objtypes.keys():
                        if objtypes[obj]==function["component_name"].rsplit(".",1)[0] and argumentcheck(objtypes,function["arguments"],arguments):
                            # print("argcheck",argumentcheck(objtypes,function["arguments"],arguments))
                            output.append(function)
                            # print(line,line.split("."+function["functionname"]+"(")[1])
                            # if "." not in line.split("."+function["functionname"]+"(")[1].split(")")[0].split("(")[0]:
                            #     print
                            #     objtypes[]
                        # else:
                        #     print(filename,function["component_name"],line,obj)
                            
                    elif function["component_name"].rsplit(".",1)[0]+"."+function["functionname"]+"(" in line and argumentcheck(objtypes,function["arguments"],arguments):
                        # print("argcheck",argumentcheck(objtypes,function["arguments"],arguments))
                        output.append(function)                    
                    # else:
                    #     print(line)
                elif " "+function["functionname"]+"(" in (" "+line.strip()):
                    if filename==function["component_name"] and argumentcheck(objtypes,function["arguments"],arguments):
                        # print("argcheck",argumentcheck(objtypes,function["arguments"],arguments))
                        output.append(function)      
    
    output = [i for n, i in enumerate(output) if i not in output[n + 1:]]
    tempoutput = copy.deepcopy(output)
    for item in [item for item in tempoutput if item["component_name"]==filename]:
        for item2 in tempoutput:
            if item2["component_name"]!=filename and item2["functionname"]==item["functionname"] and item2["noa"]==item["noa"] :
                output.remove(item2)
                # print(item2["functionname"],item2["component_name"],filename)
    return output

def getallcalls(filespath):
    output = []
    allfilefunctions = getallfilefunctions(filespath)
    for filepath in getallfiles(filespath,[".java"]):# iterate through every file
        # print(filepath.split("\\")[-1])
        Lines = open(filepath).readlines()  
        Lines = removecommentlines(Lines,"java")
        importfiles = []
        importfiles = importfiles + [item.split("\\")[-1] for item in getallfiles(filepath.rsplit("\\", 1)[0], [".java"])]
        for line in Lines:  # iterate through lines of file          
            if line.strip().startswith("import ") and line.strip().endswith(";") and not (line.strip().startswith("import java.") or line.strip().startswith("import javax.")):
                path = line.split("import ",1)[1].split(";")[0]
                if path.endswith("*"):
                    path = path.replace("*", "")
                    importfiles = importfiles + [item.split("\\")[-1] for item in
                                                 getallfiles(remainingfilepath + path.replace(".", "\\"),[".java"])]
                else:
                    importfiles.append(path.rsplit(".", 1)[-1]+".java")
                    
        importfilefunctions = []
        for importfile in importfiles:
            temp = allfilefunctions[importfile]
            for tempitem in temp:
                tempitem["component_name"] = importfile
                importfilefunctions.append(copy.deepcopy(tempitem))
            
        for functionname in allfilefunctions[filepath.split("\\")[-1]]:
            functioninfile = copy.deepcopy(functionname)
            functioninfile["component_name"] = filepath.split("\\")[-1]
            functioninfile["allcalls"] = getfunctioncalls(functioninfile,importfiles,importfilefunctions)
            
            output.append(functioninfile.copy())
    return output

def getallnodeslinks(node,allcalls):
    nodes = []
    links = []
    endnodes=[] 
    
    functioncalls = node["allcalls"]
    source = "p_"+node["label"]+"["+node["component_name"]+"]"    
    for endnode in functioncalls:
        # print(endnode)      
        if endnode["functionname"].lower().endswith(".jsp"):
            endnodename = endnode["functionname"]
            endfilename = "(JAVA_SERVER_PAGE)"
        else:
            endnodename = endnode["label"]
            endfilename = "["+endnode["component_name"]+"]"
        
        if endnodename.lower().endswith(".jsp"):

            links.append({"source" : source,
                          "target" : "p_"+endnodename.rsplit(".",1)[0]+endfilename,
                          "label"  : "External_Program"})
            nodes.append({"id"     : "p_"+endnodename.rsplit(".",1)[0]+endfilename,
                          "label"  :      endnodename.rsplit(".",1)[0]+endfilename}) 
            
        elif endnode["component_name"]!=node["component_name"]:
         
            links.append({"source" : source,
                          "target" : "p_"+endnodename+endfilename,
                          "label"  : "External_Program"})
            nodes.append({"id"     : "p_"+endnodename+endfilename,
                          "label"  :      endnodename+endfilename})
            endnodes.append(endnode)       

        else:
          
            links.append({"source" : source,
                          "target" : "p_"+endnodename+endfilename})
            nodes.append({"id"     : "p_"+endnodename+endfilename,
                          "label"  :      endnodename+endfilename})
            endnodes.append(endnode)                 
    
    if endnodes!=[]:
         for endnode in endnodes:
             if not (endnode["label"]==node["label"] and endnode["component_name"]==node["component_name"]):
                 # print(endnode)
                 endfunctionnode = [item for item in allcalls if item["component_name"] == endnode["component_name"] and item["functionname"] == endnode["functionname"] and item["noa"] == endnode["noa"]][0]
                 nodes2,links2 = getallnodeslinks(endfunctionnode,allcalls)    
                 nodes = nodes + nodes2
                 links = links + links2
                 
    nodes = [i for n, i in enumerate(nodes) if i not in nodes[n + 1:]]   
    links = [i for n, i in enumerate(links) if i not in links[n + 1:]]          
    return nodes,links

def getallreports(filespath):
    finaloutput = []
    allcalls = getallcalls(filespath)
    alltriggers,remaintriggers = geteventtriggers(filespath)
    for eventjson in alltriggers:# iterate through every file
        startnode = [item for item in allcalls if item["component_name"] == eventjson["called_name"] and item["functionname"] == eventjson["eventname"]][0]
        tempdict = copy.deepcopy(modaljson)
        tempdict["component_name"] = startnode["component_name"]
        tempdict["application"] = eventjson["application"]
        tempdict["event_name"] = startnode["functionname"]
        tempdict["nodes"]=[]
        label = startnode["label"]+"["+startnode["component_name"]+"]"
        tempdict["nodes"].append({"id"    : "p_"+label,
                                  "label" :      label})
        nodes2,tempdict["links"] = getallnodeslinks(startnode,allcalls)
        tempdict["nodes"] = tempdict["nodes"] + nodes2
        finaloutput.append(copy.deepcopy(tempdict))   
        
    for eventjson in remaintriggers:
        event = eventjson["eventname"]
        # print(eventjson["called_name"],event)
        tempdict = copy.deepcopy(modaljson)
        tempdict["component_name"] = eventjson["called_name"]
        tempdict["application"] = "UNKNOWN"
        tempdict["event_name"] = event
        tempdict["nodes"]=[{"id"    : "p_"+event,
                            "label" :      event}]
        tempdict["links"] =[]
        finaloutput.append(copy.deepcopy(tempdict))        
    return finaloutput

def dbinsertfunction(filespath,dbname,collectionname):
    col = client[dbname][collectionname]
    output = getallreports(filespath)    
    if output!=[]:   
        if col.count_documents({}) != 0 :
            col.drop()  
            print("Deleted the old",dbname,collectionname,"collection")
        col.insert_many(output)
        print("Inserted the list of jsons of",dbname,collectionname)
    else:
        print("There are no jsons in the output to insert in the DB",dbname,collectionname)

       

if __name__ == '__main__':  
    # output =  getallreports(filespath)
    # if not os.path.exists("outputs//"):
    #     os.makedirs("outputs//")
    # json.dump(output , open('outputs\\process_flow.json', 'w'), indent=4)
    # pd.DataFrame(output).to_excel("outputs\\process_flow.xlsx", index=False)     
    dbinsertfunction(filespath,dbname,collectionname)